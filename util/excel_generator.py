import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List
from source.models.Sach import Sach
from source.models.NhanVien import NhanVien
from source.models.NguonNhapSach import NguonNhapSach
from source.models.NhaPhanPhoi import NhaPhanPhoi

class ExcelGenerator:
    def generate_books_excel(self, books: List[Sach], file_path: str):
        try:
            # Tạo workbook và worksheet mới
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Danh Sách Sách"

            # Định nghĩa tiêu đề cột
            headers = ["Mã Sách", "Tên Sách", "Tác Giả", "Thể Loại", "Năm XB", "NXB", "Số Lượng", "Giá (VNĐ)"]
            ws.append(headers)

            # Style cho dòng tiêu đề
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="A94F8B", end_color="A94F8B", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Thêm dữ liệu sách vào các dòng
            for book in books:
                ws.append([
                    book.ID_Sach,
                    book.TenSach,
                    book.TacGia,
                    book.TheLoai,
                    book.NamXuatBan,
                    book.NhaXuatBan,
                    book.SoLuong,
                    book.Gia
                ])

            # Tự động điều chỉnh độ rộng cột
            column_widths = {'A': 12, 'B': 45, 'C': 25, 'D': 20, 'E': 10, 'F': 30, 'G': 10, 'H': 15}
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # Định dạng cột giá tiền
            for cell in ws['H']:
                cell.number_format = '#,##0'

            # Lưu file workbook
            wb.save(file_path)
            wb.close() # Đóng workbook để giải phóng bộ nhớ
            return True
        except Exception as e:
            print(f"Lỗi khi tạo file Excel: {e}")
            return False

    def generate_employees_excel(self, employees: List[NhanVien], file_path: str):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Danh Sách Nhân Viên"

            headers = ["Mã NV", "Họ Tên", "Giới Tính", "Chức Vụ", "Số Điện Thoại", "Email", "Trạng Thái"]
            ws.append(headers)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="5D5FEF", end_color="5D5FEF", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            for emp in employees:
                ws.append([
                    emp.ID_NhanVien, emp.HoTen, emp.GioiTinh, emp.ChucVu,
                    emp.SoDienThoai, emp.Email, emp.TrangThaiNhanVien
                ])

            column_widths = {'A': 12, 'B': 25, 'C': 12, 'D': 20, 'E': 15, 'F': 30, 'G': 15}
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            wb.save(file_path)
            wb.close()
            return True
        except Exception as e:
            print(f"Lỗi khi tạo file Excel cho nhân viên: {e}")
            return False

    def generate_suppliers_excel(self, suppliers: List[NguonNhapSach], file_path: str):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Danh Sách Nhà Cung Cấp"

            headers = ["Mã NCC", "Tên Cơ Sở", "Hình Thức Nhập", "Địa Chỉ", "Số Điện Thoại", "Email", "Trạng Thái"]
            ws.append(headers)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="6D5A72", end_color="6D5A72", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            for sup in suppliers:
                ws.append([
                    sup.ID_NguonNhap, sup.TenCoSo, sup.HinhThucNhap, sup.DiaChi,
                    sup.SoDienThoai, sup.Email, sup.TrangThaiNCC
                ])

            column_widths = {'A': 12, 'B': 30, 'C': 20, 'D': 40, 'E': 15, 'F': 30, 'G': 15}
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            wb.save(file_path)
            wb.close()
            return True
        except Exception as e:
            print(f"Lỗi khi tạo file Excel cho nhà cung cấp: {e}")
            return False

    def generate_distributors_excel(self, distributors: List[NhaPhanPhoi], file_path: str):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Danh Sách Nhà Phân Phối"

            headers = ["Mã NPP", "Tên Cơ Sở", "Địa Chỉ", "Số Điện Thoại", "Email", "Trạng Thái"]
            ws.append(headers)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="C08261", end_color="C08261", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            for dist in distributors:
                ws.append([
                    dist.ID_NguonXuat, dist.TenCoSo, dist.DiaChi,
                    dist.SoDienThoai, dist.Email, dist.TrangThaiNPP
                ])

            column_widths = {'A': 12, 'B': 30, 'C': 40, 'D': 15, 'E': 30, 'F': 15}
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            wb.save(file_path)
            wb.close()
            return True
        except Exception as e:
            print(f"Lỗi khi tạo file Excel cho nhà phân phối: {e}")
            return False